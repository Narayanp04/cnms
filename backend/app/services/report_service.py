"""ConnectXperts NMS - Report Generation Service"""
import logging
import io
import csv
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Dict
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    Spacer, PageBreak, Image
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
import xlsxwriter

from app.database import AsyncSessionLocal
from app.models.device import Device
from app.models.ping_result import PingResult, PingStatus
from app.models.sla_report import SLAReport, SLAReportPeriod

logger = logging.getLogger(__name__)


class ReportService:
    """Service for generating reports in various formats."""
    
    async def generate_daily_report(self, customer_id: Optional[int] = None) -> Dict:
        """Generate daily summary report."""
        return await self._generate_periodic_report("daily", timedelta(days=1), customer_id)
    
    async def generate_weekly_report(self, customer_id: Optional[int] = None) -> Dict:
        """Generate weekly summary report."""
        return await self._generate_periodic_report("weekly", timedelta(weeks=1), customer_id)
    
    async def generate_monthly_report(self, customer_id: Optional[int] = None) -> Dict:
        """Generate monthly summary report."""
        return await self._generate_periodic_report("monthly", timedelta(days=30), customer_id)
    
    async def generate_yearly_report(self, customer_id: Optional[int] = None) -> Dict:
        """Generate yearly summary report."""
        return await self._generate_periodic_report("yearly", timedelta(days=365), customer_id)
    
    async def _generate_periodic_report(self, period: str, delta: timedelta, customer_id: Optional[int] = None) -> Dict:
        """Generate a periodic report with device summaries."""
        now = datetime.now(timezone.utc)
        start = now - delta
        
        async with AsyncSessionLocal() as db:
            from sqlalchemy import select, func, case
            
            # Get devices
            query = select(Device).where(Device.is_deleted == False)
            if customer_id:
                query = query.where(Device.customer_id == customer_id)
            
            result = await db.execute(query)
            devices = result.scalars().all()
            
            report_data = {
                "period": period,
                "generated_at": now.isoformat(),
                "period_start": start.isoformat(),
                "period_end": now.isoformat(),
                "total_devices": len(devices),
                "devices": []
            }
            
            for device in devices:
                ping_result = await db.execute(
                    select(
                        func.count(PingResult.id).label('total'),
                        func.sum(
                            case((PingResult.status == PingStatus.SUCCESS, 1), else_=0)
                        ).label('successful'),
                        func.avg(PingResult.latency_ms).label('avg_latency'),
                        func.max(PingResult.latency_ms).label('max_latency'),
                        func.min(PingResult.latency_ms).label('min_latency'),
                        func.avg(PingResult.packet_loss_percent).label('avg_packet_loss'),
                    ).where(
                        PingResult.device_id == device.id,
                        PingResult.timestamp >= start,
                        PingResult.timestamp <= now
                    )
                )
                stats = ping_result.one()
                
                total = stats.total or 0
                successful = stats.successful or 0
                availability = (successful / total * 100) if total > 0 else 100.0
                
                device_data = {
                    "id": device.id,
                    "hostname": device.hostname,
                    "ip_address": device.ip_address,
                    "customer_name": device.customer_name,
                    "site_name": device.site_name,
                    "status": device.status.value,
                    "availability": round(availability, 2),
                    "avg_latency_ms": round(float(stats.avg_latency), 2) if stats.avg_latency else None,
                    "max_latency_ms": round(float(stats.max_latency), 2) if stats.max_latency else None,
                    "avg_packet_loss": round(float(stats.avg_packet_loss), 2) if stats.avg_packet_loss else 0,
                    "total_pings": total,
                    "failed_pings": total - int(successful or 0)
                }
                report_data["devices"].append(device_data)
            
            return report_data
    
    def export_to_pdf(self, report_data: Dict, title: str = "CNMS Report") -> bytes:
        """Export report data to PDF format."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            spaceAfter=30
        )
        elements.append(Paragraph(f"ConnectXperts NMS Report", title_style))
        elements.append(Paragraph(f"{title}", styles['Heading1']))
        elements.append(Spacer(1, 12))
        
        # Report info
        info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10)
        elements.append(Paragraph(f"Period: {report_data.get('period_start', 'N/A').split('T')[0]} to {report_data.get('period_end', 'N/A').split('T')[0]}", info_style))
        elements.append(Paragraph(f"Generated: {report_data.get('generated_at', 'N/A').split('T')[0]}", info_style))
        elements.append(Paragraph(f"Total Devices: {report_data.get('total_devices', 0)}", info_style))
        elements.append(Spacer(1, 20))
        
        # Summary Table
        devices = report_data.get('devices', [])
        if devices:
            table_data = [['Hostname', 'IP Address', 'Status', 'Availability %', 'Avg Latency', 'Packet Loss']]
            for d in devices:
                table_data.append([
                    d.get('hostname', ''),
                    d.get('ip_address', ''),
                    d.get('status', ''),
                    f"{d.get('availability', 0):.2f}%",
                    f"{d.get('avg_latency_ms', 'N/A')} ms",
                    f"{d.get('avg_packet_loss', 0):.2f}%"
                ])
            
            table = Table(table_data, colWidths=[1.5*inch, 1.2*inch, 0.8*inch, 1*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
    
    def export_to_excel(self, report_data: Dict) -> bytes:
        """Export report data to Excel format."""
        output = io.BytesIO()
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Report"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Title
        sheet.merge_cells('A1:F1')
        title_cell = sheet['A1']
        title_cell.value = f"ConnectXperts NMS Report"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ['Hostname', 'IP Address', 'Customer', 'Site', 'Status', 'Availability %', 'Avg Latency (ms)', 'Packet Loss %']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data
        for row, device in enumerate(report_data.get('devices', []), 4):
            sheet.cell(row=row, column=1, value=device.get('hostname', ''))
            sheet.cell(row=row, column=2, value=device.get('ip_address', ''))
            sheet.cell(row=row, column=3, value=device.get('customer_name', ''))
            sheet.cell(row=row, column=4, value=device.get('site_name', ''))
            sheet.cell(row=row, column=5, value=device.get('status', ''))
            sheet.cell(row=row, column=6, value=device.get('availability', 0))
            sheet.cell(row=row, column=7, value=device.get('avg_latency_ms', ''))
            sheet.cell(row=row, column=8, value=device.get('avg_packet_loss', 0))
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
    
    def export_to_csv(self, report_data: Dict) -> str:
        """Export report data to CSV format."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['ConnectXperts NMS Report'])
        writer.writerow([f"Period: {report_data.get('period_start', 'N/A')} to {report_data.get('period_end', 'N/A')}"])
        writer.writerow([])
        
        # Column headers
        writer.writerow(['Hostname', 'IP Address', 'Customer', 'Site', 'Status', 'Availability %', 'Avg Latency (ms)', 'Packet Loss %', 'Total Pings', 'Failed Pings'])
        
        # Data
        for device in report_data.get('devices', []):
            writer.writerow([
                device.get('hostname', ''),
                device.get('ip_address', ''),
                device.get('customer_name', ''),
                device.get('site_name', ''),
                device.get('status', ''),
                device.get('availability', 0),
                device.get('avg_latency_ms', ''),
                device.get('avg_packet_loss', 0),
                device.get('total_pings', 0),
                device.get('failed_pings', 0)
            ])
        
        return output.getvalue()
